from __future__ import annotations
import sys
import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
from pathlib import Path
import openpyxl as opxl
from datetime import datetime
import csv
import re

BG_DARK = '#1a1a2e'
BG_CARD = '#16213e'
BG_INPUT = '#0f3460'
FG_TEXT = '#e0e0e0'
FG_LABEL = '#a8b2d1'
FG_TITLE = '#ffffff'
ACCENT = '#e94560'
ACCENT_HOVER = '#ff6b81'
SUCCESS = '#00d2d3'
WARNING = '#feca57'
BORDER = '#233554'


class StockReportApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title('BLP Stock Report Generator')
        self.root.geometry('900x1000')
        self.root.minsize(850, 900)
        self.root.configure(bg=BG_DARK)

        self.var_source = tk.StringVar()
        self.var_output = tk.StringVar()
        self.var_selected_sheets = []
        self.checkboxes = []
        self.processing = False
        self.sheet_files = {}
        self.input_file_vars = {}

        self._build_ui()

    def _build_ui(self):
        # Main scrollable container
        self.main_canvas = tk.Canvas(self.root, bg=BG_DARK, highlightthickness=0)
        self.main_scroll = tk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        self.main_canvas.configure(yscrollcommand=self.main_scroll.set)
        
        self.main_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        container = tk.Frame(self.main_canvas, bg=BG_DARK, padx=20, pady=15)
        self.main_canvas.create_window((0, 0), window=container, anchor='nw')
        
        container.bind("<Configure>", lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all")))

        # Title
        title_frame = tk.Frame(container, bg=BG_DARK)
        title_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(title_frame, text='📊', font=('Segoe UI Emoji', 24),
                 bg=BG_DARK, fg=FG_TITLE).pack(side=tk.LEFT, padx=(0, 8))
        title_text = tk.Frame(title_frame, bg=BG_DARK)
        title_text.pack(side=tk.LEFT)
        tk.Label(title_text, text='BLP Stock Report Generator',
                 font=('Segoe UI', 16, 'bold'), bg=BG_DARK, fg=FG_TITLE).pack(anchor='w')
        tk.Label(title_text, text='Auto-match file ke sheet | FEFO Sales | Proses Semua Sheet Sekaligus',
                 font=('Segoe UI', 9), bg=BG_DARK, fg=FG_LABEL).pack(anchor='w')

        # File Utama
        self._section_label(container, '📁 File Utama (Wajib)')
        card1 = self._card(container)
        self._file_row(card1, 'File Sumber (.xlsx)', self.var_source, 0)
        self._file_row(card1, 'File Output (opsional)', self.var_output, 1, save=True)

        tk.Button(card1, text='📋 Load Sheets', font=('Segoe UI', 9),
                  bg=BG_INPUT, fg=FG_TEXT, relief='flat', cursor='hand2',
                  command=self._load_sheets).grid(row=2, column=0, sticky='w', padx=8, pady=4)

        # File Input per Sheet
        self._section_label(container, '📂 File Input per Sheet (Pilih Satu-Satu)')
        self.input_card = self._card(container)
        
        self.input_files_container = tk.Frame(self.input_card, bg=BG_CARD)
        self.input_files_container.pack(fill=tk.X, padx=5, pady=5)
        
        tk.Label(self.input_card, text='(Klik SM/Odoo utk pilih file, lalu Load Sheets dulu)', font=('Segoe UI', 8),
                 bg=BG_CARD, fg=FG_LABEL).pack(anchor='w', padx=8, pady=(0, 5))

        # Sheet selection
        self._section_label(container, '📋 Pilih Sheet yang Diproses')
        sheet_card = self._card(container)
        sheet_frame = tk.Frame(sheet_card, bg=BG_CARD)
        sheet_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.sheet_canvas = tk.Canvas(sheet_frame, bg=BG_CARD, height=150, highlightthickness=0)
        self.sheet_scroll = tk.Scrollbar(sheet_frame, orient="vertical", command=self.sheet_canvas.yview)
        self.sheet_inner = tk.Frame(self.sheet_canvas, bg=BG_CARD)
        self.sheet_canvas.create_window((0, 0), window=self.sheet_inner, anchor='nw')
        self.sheet_canvas.configure(yscrollcommand=self.sheet_scroll.set)
        
        self.sheet_inner.bind("<Configure>", lambda e: self.sheet_canvas.configure(scrollregion=self.sheet_canvas.bbox("all")))
        
        self.sheet_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.sheet_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        tk.Label(sheet_card, text='(Centang sheet yang mau diproses, kosongkan = semua)', font=('Segoe UI', 8),
                 bg=BG_CARD, fg=FG_LABEL).pack(anchor='w', padx=8, pady=(0, 5))

        # Process button
        btn_frame = tk.Frame(container, bg=BG_DARK)
        btn_frame.pack(fill=tk.X, pady=(10, 5))

        self.btn_process = tk.Button(
            btn_frame, text='🚀  PROSES', font=('Segoe UI', 11, 'bold'),
            bg=ACCENT, fg='white', activebackground=ACCENT_HOVER, activeforeground='white',
            relief='flat', cursor='hand2', padx=30, pady=8,
            command=self._on_process
        )
        self.btn_process.pack(expand=True)

        # Progress
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Custom.Horizontal.TProgressbar', background=SUCCESS, troughcolor=BG_INPUT)
        self.progress = ttk.Progressbar(container, style='Custom.Horizontal.TProgressbar', mode='indeterminate', length=400)
        self.progress.pack(fill=tk.X, pady=(5, 5))

        # Log
        self._section_label(container, '📋 Log Output')
        self.log = scrolledtext.ScrolledText(container, height=8, font=('Consolas', 9),
                                              bg=BG_INPUT, fg=FG_TEXT, insertbackground=FG_TEXT,
                                              relief='flat', bd=0, wrap=tk.WORD, state=tk.DISABLED)
        self.log.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log.tag_configure('error', foreground='#ff6b6b')
        self.log.tag_configure('success', foreground=SUCCESS)
        self.log.tag_configure('warn', foreground=WARNING)
        self.log.tag_configure('info', foreground=FG_LABEL)

    def _section_label(self, parent, text):
        tk.Label(parent, text=text, font=('Segoe UI', 10, 'bold'),
                 bg=BG_DARK, fg=FG_TITLE, anchor='w').pack(fill=tk.X, pady=(10, 3))

    def _card(self, parent):
        card = tk.Frame(parent, bg=BG_CARD, padx=10, pady=8, highlightbackground=BORDER, highlightthickness=1)
        card.pack(fill=tk.X, pady=(0, 2))
        return card

    def _file_row(self, parent, label, var, row, save=False):
        tk.Label(parent, text=label + ':', font=('Segoe UI', 9),
                 bg=BG_CARD, fg=FG_LABEL, width=20, anchor='w').grid(row=row, column=0, sticky='w', padx=8, pady=4)
        entry = tk.Entry(parent, textvariable=var, font=('Segoe UI', 9),
                        bg=BG_INPUT, fg=FG_TEXT, insertbackground=FG_TEXT, relief='flat', bd=0)
        entry.grid(row=row, column=1, sticky='ew', padx=5, pady=4, ipady=4)
        cmd = self._make_save_dialog(var) if save else self._make_open_dialog(var)
        tk.Button(parent, text='📂', font=('Segoe UI Emoji', 10), bg=BG_CARD, fg=FG_TEXT,
                 relief='flat', cursor='hand2', activebackground=BG_INPUT, command=cmd).grid(row=row, column=2, padx=5, pady=4)

    def _make_open_dialog(self, var):
        def handler():
            path = filedialog.askopenfilename(filetypes=[('Excel/CSV', '*.xlsx *.xls *.csv'), ('All', '*.*')])
            if path: var.set(path)
        return handler

    def _make_save_dialog(self, var):
        def handler():
            path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel', '*.xlsx'), ('All', '*.*')])
            if path: var.set(path)
        return handler

    def _browse_folder(self):
        path = filedialog.askdirectory()
        if path:
            self.var_folder.set(path)
            self._log(f'Folder dipilih: {path}', 'info')
            if self.var_source.get():
                self._match_files_to_sheets()

    def _load_sheets(self):
        source = self.var_source.get().strip().strip('"')
        if not source:
            messagebox.showwarning('Peringatan', 'Pilih file sumber terlebih dahulu!')
            return
        src = Path(source).expanduser().resolve()
        if not src.exists():
            messagebox.showerror('Error', 'File tidak ditemukan!')
            return
        try:
            wb = opxl.load_workbook(src, read_only=True)
            for widget in self.sheet_inner.winfo_children(): widget.destroy()
            self.checkboxes = []
            for sheet_name in wb.sheetnames:
                var = tk.BooleanVar()
                cb = tk.Checkbutton(self.sheet_inner, text=sheet_name, variable=var,
                                  font=('Segoe UI', 9), bg=BG_CARD, fg=FG_TEXT,
                                  selectcolor=BG_INPUT, anchor='w')
                cb.pack(fill=tk.X)
                self.checkboxes.append((sheet_name, var))
            self._build_input_files_ui()
            self._log(f'Loaded {len(wb.sheetnames)} sheets', 'success')
            wb.close()
        except Exception as e:
            messagebox.showerror('Error', f'Gagal load sheets: {str(e)}')

    def _build_input_files_ui(self):
        for widget in self.input_files_container.winfo_children(): widget.destroy()
        
        self.input_file_vars = {}
        
        for sheet_name, var in self.checkboxes:
            sheet_frame = tk.Frame(self.input_files_container, bg=BG_DARK, pady=3)
            sheet_frame.pack(fill=tk.X, padx=5)
            
            tk.Label(sheet_frame, text=f'{sheet_name}:', font=('Segoe UI', 9, 'bold'),
                     bg=BG_DARK, fg=FG_TEXT, width=18, anchor='w').pack(side=tk.LEFT)
            
            sm_var = tk.StringVar()
            odoo_var = tk.StringVar()
            self.input_file_vars[sheet_name] = {'stock_masuk': sm_var, 'odoo': odoo_var}
            
            tk.Button(sheet_frame, text='SM', font=('Segoe UI', 8),
                      bg=BG_INPUT, fg=FG_TEXT, relief='flat', cursor='hand2',
                      command=lambda v=sm_var: self._select_input_file(v, 'Stock Masuk')).pack(side=tk.LEFT, padx=2)
            
            tk.Button(sheet_frame, text='Odoo', font=('Segoe UI', 8),
                      bg=BG_INPUT, fg=FG_TEXT, relief='flat', cursor='hand2',
                      command=lambda v=odoo_var: self._select_input_file(v, 'Odoo')).pack(side=tk.LEFT, padx=2)
            
            tk.Label(sheet_frame, text='(belum dipilih)', font=('Segoe UI', 8),
                     bg=BG_DARK, fg=FG_LABEL, anchor='w').pack(side=tk.LEFT, padx=5)

    def _select_input_file(self, var, file_type):
        path = filedialog.askopenfilename(filetypes=[('Excel/CSV', '*.xlsx *.xls *.csv'), ('All', '*.*')])
        if path:
            var.set(path)
            fname = Path(path).name
            for widget in self.input_files_container.winfo_children():
                for child in widget.winfo_children():
                    if hasattr(child, 'configure'):
                        try:
                            if child.cget('text') == '(belum dipilih)':
                                child.configure(text=fname[:20], fg=SUCCESS)
                        except: pass

    def _match_files_to_sheets(self):
        folder = self.var_folder.get().strip()
        if not folder: return
        
        folder_path = Path(folder)
        files = list(folder_path.glob('*.xlsx')) + list(folder_path.glob('*.csv'))
        
        self.sheet_files = {}
        for sheet_name, var in self.checkboxes:
            self.sheet_files[sheet_name] = {'stock_masuk': None, 'odoo': None, 'commercial': None}
        
        matched = []
        for f in files:
            fname_lower = f.stem.lower().replace('_', ' ').replace('-', ' ')
            for sheet_name, var in self.checkboxes:
                sheet_clean = sheet_name.lower().replace('_', ' ')
                if sheet_clean in fname_lower or sheet_clean.replace(' ', '_') in fname_lower:
                    if 'stock masuk' in fname_lower or 'stock_masuk' in f.stem.lower():
                        self.sheet_files[sheet_name]['stock_masuk'] = f
                        matched.append(f'{sheet_name}: Stock Masuk ✓')
                    elif fname_lower.startswith('odoo'):
                        self.sheet_files[sheet_name]['odoo'] = f
                        matched.append(f'{sheet_name}: Odoo ✓')
                    elif 'commercial' in fname_lower or 'fruition' in fname_lower:
                        self.sheet_files[sheet_name]['commercial'] = f
        
        self._log(f'File matched: {len(matched)} file', 'success')
        for m in matched[:5]:
            self._log(f'  {m}', 'info')
        if len(matched) > 5:
            self._log(f'  ... dan {len(matched)-5} lagi', 'info')

    def _log(self, text, tag=None):
        self.log.configure(state=tk.NORMAL)
        if tag: self.log.insert(tk.END, text + '\n', tag)
        else: self.log.insert(tk.END, text + '\n')
        self.log.see(tk.END)
        self.log.configure(state=tk.DISABLED)

    def _clear_log(self):
        self.log.configure(state=tk.NORMAL)
        self.log.delete(1.0, tk.END)
        self.log.configure(state=tk.DISABLED)

    def _get_selected_sheets(self):
        return [sn for sn, v in self.checkboxes if v.get()]

    def _on_process(self):
        if self.processing: return
        source = self.var_source.get().strip().strip('"')
        if not source: self._log('[ERROR] File sumber harus diisi!', 'error'); return
        
        src = Path(source).expanduser().resolve()
        if not src.exists(): self._log(f'[ERROR] File tidak ditemukan: {src}', 'error'); return
        
        self.processing = True
        self.btn_process.configure(state=tk.DISABLED, text='⏳  MEMPROSES...')
        self.progress.start(15)
        self._clear_log()
        thread = threading.Thread(target=self._run_process, args=(src,), daemon=True)
        thread.start()

    def _finish_process(self):
        self.processing = False
        self.progress.stop()
        self.btn_process.configure(state=tk.NORMAL, text='🚀  PROSES')

    def _run_process(self, src: Path):
        try:
            self._process(src)
        except Exception as e:
            self.root.after(0, self._log, f'[ERROR] {e}', 'error')
        finally:
            self.root.after(0, self._finish_process)

    def _log_safe(self, text, tag=None):
        self.root.after(0, self._log, text, tag)

    def _process(self, src: Path):
        from datetime import datetime
        import csv

        def norm(text):
            if text is None: return ''
            return str(text).strip().lower().replace('\n', ' ').replace('  ', ' ')

        def safe_float(value):
            if value is None or value == '': return 0.0
            try: return float(value)
            except: return 0.0

        def header_map(ws, row=2):
            mapping = {}
            for col in range(1, ws.max_column + 1):
                key = norm(ws.cell(row, col).value)
                if key: mapping[key] = col
            return mapping

        def find_header_col(headers, *candidates):
            for c in candidates:
                col = headers.get(c)
                if col is not None: return col
            return None

        def first_nonblank(*vals):
            for v in vals:
                if v not in (None, ''): return v
            return None

        def is_branch_sheet(ws):
            headers = header_map(ws)
            return 'stock awal' in headers and 'stock akhir' in headers

        def guess_output_name(src):
            MONTHS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            MONTH_RE = re.compile(r'\b(' + '|'.join(MONTHS) + r')\b', re.IGNORECASE)
            YEAR_RE = re.compile(r'\b(20\d{2})\b')
            stem = src.stem
            m = MONTH_RE.search(stem)
            y = YEAR_RE.search(stem)
            if m:
                idx = MONTHS.index(m.group(1).lower())
                next_idx = (idx + 1) % 12
                next_year = int(y.group(1)) + (1 if next_idx == 0 and idx == 11 else 0) if y else None
                new_stem = MONTH_RE.sub(m[:1].upper() + m[1:].lower(), stem, count=1)
                if y and next_year: new_stem = YEAR_RE.sub(str(next_year), new_stem)
                return src.with_name(new_stem + src.suffix)
            return src.with_name(src.stem + ' - bulan baru' + src.suffix)

        def load_odoo_transfer(path):
            wb = opxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb['Sheet1']
            results = []
            for row in range(2, ws.max_row + 1):
                sku = ws.cell(row, 9).value
                sku_batch = ws.cell(row, 10).value
                lot = ws.cell(row, 11).value
                product = ws.cell(row, 12).value
                qty_done = ws.cell(row, 13).value
                qty = ws.cell(row, 14).value
                if product and lot:
                    qty = safe_float(qty_done) if qty_done and qty_done > 0 else safe_float(qty)
                    if qty > 0:
                        results.append({'sku': sku, 'sku_batch': sku_batch, 'lot': lot, 'product': product, 'qty': qty})
            wb.close()
            return results

        def process_transfer(ws, headers, data, col_name, existing_max_row):
            product_col = find_header_col(headers, 'product')
            lot_col = find_header_col(headers, 'lot/serial number')
            target_col = find_header_col(headers, col_name, f'{col_name} ')
            if not all([product_col, lot_col, target_col]): return 0, 0
            
            updated, added, new_row = 0, 0, existing_max_row + 1
            for item in data:
                product, lot, qty = str(item['product']).strip(), str(item['lot']).strip(), item['qty']
                found = False
                for row in range(3, existing_max_row + 1):
                    if str(ws.cell(row, product_col) or '').strip() == product and str(ws.cell(row, lot_col) or '').strip() == lot:
                        ws.cell(row, target_col).value = (ws.cell(row, target_col).value or 0) + qty
                        updated += 1
                        found = True
                        break
                if not found:
                    ws.cell(new_row, 1).value = item.get('sku', '')
                    ws.cell(new_row, 2).value = item.get('sku_batch', '')
                    ws.cell(new_row, 3).value = item['product']
                    ws.cell(new_row, 4).value = lot
                    ws.cell(new_row, 7).value = 0
                    ws.cell(new_row, 10).value = 0
                    ws.cell(new_row, 11).value = 0
                    ws.cell(new_row, 13).value = 0
                    ws.cell(new_row, target_col).value = qty
                    new_row += 1
                    added += 1
            return updated, added

        def add_formulas(ws, start_row, end_row):
            sum_cols = {7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 16: 'P', 17: 'Q'}
            for col_num, col_letter in sum_cols.items():
                ws.cell(1, col_num).value = f'=SUM({col_letter}{start_row}:{col_letter}{end_row})'
            for row in range(start_row, end_row + 1):
                r = str(row)
                ws.cell(row, 1).value = f'=MID(C{r},FIND("[",C{r})+1,FIND("]",C{r})-FIND("[",C{r})-1)'
                ws.cell(row, 2).value = f'=A{r}&D{r}'
                ws.cell(row, 5).value = f'=IF(ISBLANK(F{r}),"",DATEDIF(MIN(TODAY()),F{r}),MAX(TODAY()),F{r}),"M")*IF(F{r}>=TODAY(),1,-1))'
                ws.cell(row, 9).value = f'=SUM(G{r}:H{r})'
                ws.cell(row, 12).value = f'=SUM(I{r}:J{r})-K{r}'
                ws.cell(row, 14).value = f'=L{r}-M{r}'
                ws.cell(row, 17).value = f'=P{r}-M{r}'

        target_sheets = self._get_selected_sheets()
        output_str = self.var_output.get().strip().strip('"')
        out = Path(output_str).expanduser().resolve() if output_str else guess_output_name(src)

        self._log_safe('=' * 55)
        self._log_safe('  PEMBUAT FILE LAPORAN STOCK BULAN BARU', 'info')
        self._log_safe('=' * 55)
        self._log_safe(f'  Sumber : {src.name}')
        self._log_safe(f'  Output : {out.name}')
        self._log_safe(f'  Sheets : {", ".join(target_sheets) if target_sheets else "SEMUA"}', 'info')

        wb = opxl.load_workbook(src, data_only=False)
        wb_values = opxl.load_workbook(src, data_only=True)
        stock_map = {}
        summary = {'sheets': 0, 'rows': 0, 'sm_filled': 0, 'odoo_filled': 0}
        RESET_COLUMNS = ['stock masuk', 'inventory adjustment', 'retur', 'sales', 'odoo']

        for sheet_name in wb.sheetnames:
            if target_sheets and sheet_name not in target_sheets: continue
            ws = wb[sheet_name]
            if not is_branch_sheet(ws): continue

            self._log_safe(f'\nMemproses: "{sheet_name}"...')
            summary['sheets'] += 1

            try:
                ws_val = wb_values[sheet_name]
                headers = header_map(ws)
                sku_col = find_header_col(headers, 'sku')
                lot_col = find_header_col(headers, 'lot/serial number')
                stock_awal = find_header_col(headers, 'stock awal', 'stock awal ')
                stock_akhir = find_header_col(headers, 'stock akhir', 'stock akhir ')
                if not all([stock_awal, stock_akhir]): continue

                existing_max = ws.max_row
                for row in range(3, ws.max_row + 1):
                    if any(ws.cell(row, c).value not in (None, '') for c in range(1, 15)):
                        closing = safe_float(ws_val.cell(row, stock_akhir).value)
                        ws.cell(row, stock_awal).value = closing
                        for cn in RESET_COLUMNS:
                            ci = headers.get(cn)
                            if ci: ws.cell(row, ci).value = 0
                        summary['rows'] += 1
                        if sku_col and lot_col:
                            sku = first_nonblank(ws_val.cell(row, sku_col).value, ws.cell(row, sku_col).value)
                            lot = first_nonblank(ws_val.cell(row, lot_col).value, ws.cell(row, lot_col).value)
                            if sku and lot:
                                stock_map[(sheet_name, str(sku), str(lot))] = closing

                # Process Stock Masuk
                vars = self.input_file_vars.get(sheet_name, {})
                sm_path = (vars.get('stock_masuk') or tk.StringVar()).get()
                if sm_path:
                    sm_p = Path(sm_path)
                    if sm_p.exists():
                        data = load_odoo_transfer(sm_p)
                        up, add = process_transfer(ws, headers, data, 'stock masuk', existing_max)
                        summary['sm_filled'] += up + add
                        existing_max = ws.max_row
                        if up or add: self._log_safe(f'  → Stock Masuk: {up} update, {add} baru', 'success')

                # Process Odoo
                odoo_path = (vars.get('odoo') or tk.StringVar()).get()
                if odoo_path:
                    o_p = Path(odoo_path)
                    if o_p.exists():
                        data = load_odoo_transfer(o_p)
                        up, add = process_transfer(ws, headers, data, 'odoo', existing_max)
                        summary['odoo_filled'] += up + add
                        existing_max = ws.max_row
                        if up or add: self._log_safe(f'  → Odoo: {up} update, {add} baru', 'success')

                add_formulas(ws, 3, ws.max_row)

            except Exception as e:
                self._log_safe(f'  [ERROR] {e}', 'error')

        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcOnSave = True
        except: pass

        wb.save(out)
        self._log_safe('\n' + '=' * 55)
        self._log_safe('  RINGKASAN', 'info')
        self._log_safe('=' * 55)
        self._log_safe(f'  Sheet diproses : {summary["sheets"]}')
        self._log_safe(f'  Baris di-reset : {summary["rows"]}')
        if summary['sm_filled']: self._log_safe(f'  Stock Masuk    : {summary["sm_filled"]}')
        if summary['odoo_filled']: self._log_safe(f'  Odoo         : {summary["odoo_filled"]}')
        self._log_safe(f'\n  ✅ File dibuat : {out.name}', 'success')
        self._log_safe('=' * 55)


def main():
    root = tk.Tk()
    app = StockReportApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()