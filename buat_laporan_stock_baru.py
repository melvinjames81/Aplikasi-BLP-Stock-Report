from __future__ import annotations
import sys
import re
import argparse
import csv
import os
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import openpyxl

MONTHS = [
    'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec'
]
MONTH_RE = re.compile(r'\b(' + '|'.join(MONTHS) + r')\b', re.IGNORECASE)
YEAR_RE = re.compile(r'\b(20\d{2})\b')

RESET_COLUMNS = [
    'stock masuk', 'inventory adjustment', 'inventory adjusment',
    'retur', 'sales', 'odoo',
]


def norm(text) -> str:
    if text is None:
        return ''
    return str(text).strip().lower().replace('\n', ' ').replace('  ', ' ')


def month_title(m: str) -> str:
    return m[:1].upper() + m[1:].lower()


def guess_output_name(src: Path) -> Path:
    stem = src.stem
    m = MONTH_RE.search(stem)
    y = YEAR_RE.search(stem)
    if m:
        current = m.group(1).lower()
        idx = MONTHS.index(current)
        next_idx = (idx + 1) % 12
        next_year = None
        if y:
            next_year = int(y.group(1)) + (1 if next_idx == 0 and idx == 11 else 0)
        new_stem = MONTH_RE.sub(month_title(MONTHS[next_idx]), stem, count=1)
        if y and next_year is not None:
            new_stem = YEAR_RE.sub(str(next_year), new_stem, count=1)
        return src.with_name(new_stem + src.suffix)
    return src.with_name(src.stem + ' - bulan baru' + src.suffix)


def header_map(ws, row=2) -> Dict[str, int]:
    mapping = {}
    for col in range(1, ws.max_column + 1):
        key = norm(ws.cell(row, col).value)
        if key:
            mapping[key] = col
    return mapping


def find_header_col(headers: Dict[str, int], *candidates: str) -> Optional[int]:
    for c in candidates:
        col = headers.get(c)
        if col is not None:
            return col
    return None


def is_branch_sheet(ws) -> bool:
    headers = header_map(ws)
    return 'stock awal' in headers and 'stock akhir' in headers


def first_nonblank(*vals):
    for v in vals:
        if v not in (None, ''):
            return v
    return None


def safe_float(value) -> float:
    if value is None or value == '':
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def parse_expiry_date(date_str) -> Optional[datetime]:
    if date_str is None or str(date_str).strip() == '':
        return None
    s = str(date_str).strip()
    for fmt in ('%b-%y', '%b-%Y', '%B-%y', '%B-%Y', '%m/%d/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def find_sheet_ci(wb, target_name: str):
    target_lower = target_name.lower()
    for name in wb.sheetnames:
        if name.lower() == target_lower:
            return wb[name]
    return None


def match_files_to_sheets(folder_path: Path, sheet_names: List[str]) -> Dict[str, Dict[str, Path]]:
    """
    Match files in folder to sheets based on naming convention.
    
    Expected patterns:
    - Stock Masuk_[SHEET_NAME].xlsx or Stock_Masuk_[SHEET_NAME].xlsx
    - Odoo_[SHEET_NAME].xlsx
    - Commercial_[SHEET_NAME].xlsx or Commercial.xlsx (shared)
    """
    files = list(folder_path.glob('*.xlsx')) + list(folder_path.glob('*.csv'))
    matched = {}
    
    # Initialize for each sheet
    for sheet_name in sheet_names:
        matched[sheet_name] = {
            'stock_masuk': None,
            'odoo': None,
            'commercial': None
        }
    
    # Also check for shared commercial file
    shared_commercial = None
    for f in files:
        fname_lower = f.stem.lower()
        if fname_lower in ['commercial', 'stock fruition', 'stock_fruition']:
            shared_commercial = f
            break
    
    # Match files to sheets
    for f in files:
        fname = f.stem.lower()
        fname_clean = fname.replace('_', ' ').replace('-', ' ')
        
        for sheet_name in sheet_names:
            sheet_clean = sheet_name.lower()
            
            # Check for Stock Masuk
            if 'stock masuk' in fname_lower or 'stock_masuk' in fname_lower:
                if sheet_clean in fname_lower or sheet_clean.replace(' ', '_') in fname_lower:
                    matched[sheet_name]['stock_masuk'] = f
                    continue
            
            # Check for Odoo
            if fname_lower.startswith('odoo'):
                if sheet_clean in fname_lower or sheet_clean.replace(' ', '_') in fname_lower:
                    matched[sheet_name]['odoo'] = f
                    continue
            
            # Check for Commercial
            if 'commercial' in fname_lower or 'fruition' in fname_lower:
                if sheet_clean in fname_lower or sheet_clean.replace(' ', '_') in fname_lower:
                    matched[sheet_name]['commercial'] = f
                    continue
    
    # Apply shared commercial to all sheets
    if shared_commercial:
        for sheet_name in sheet_names:
            if matched[sheet_name]['commercial'] is None:
                matched[sheet_name]['commercial'] = shared_commercial
    
    return matched


def load_odoo_transfer_data(path: Path) -> List[Dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['Sheet1']
    
    results = []
    for row in range(2, ws.max_row + 1):
        sku = ws.cell(row, 9).value
        sku_batch = ws.cell(row, 10).value
        lot = ws.cell(row, 11).value
        product = ws.cell(row, 12).value
        qty_done = ws.cell(row, 13).value
        qty = ws.cell(row, 14).value
        
        if product and lot:
            final_qty = safe_float(qty_done) if qty_done and qty_done > 0 else safe_float(qty)
            if final_qty > 0:
                results.append({
                    'sku': sku,
                    'sku_batch': sku_batch,
                    'lot': lot,
                    'product': product,
                    'qty': final_qty
                })
    
    wb.close()
    return results


def load_commercial_data(path: Path, sheet_name: Optional[str] = None) -> Dict[str, float]:
    ext = path.suffix.lower()
    sales_data: Dict[str, float] = {}

    if ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        
        if sheet_name:
            ws = None
            for name in wb.sheetnames:
                if name.lower().strip() == sheet_name.lower().strip():
                    ws = wb[name]
                    break
            if ws is None:
                print(f'  [WARN] Sheet "{sheet_name}" tidak ditemukan di {path.name}.')
                print(f'  Sheet tersedia: {", ".join(wb.sheetnames)}')
                wb.close()
                return sales_data
        else:
            ws = wb[wb.sheetnames[-1]]
            print(f'  Menggunakan sheet terakhir: "{wb.sheetnames[-1]}"')

        header_row = None
        for r in range(1, min(21, ws.max_row + 1)):
            for c in range(1, min(15, ws.max_column + 1)):
                val = norm(ws.cell(r, c).value)
                if val in ('sku no', 'sku no.', 'sku'):
                    header_row = r
                    break
            if header_row:
                break

        if header_row is None:
            print(f'  [WARN] Tidak bisa menemukan header SKU di {path.name}')
            wb.close()
            return sales_data

        headers = {}
        for c in range(1, ws.max_column + 1):
            key = norm(ws.cell(header_row, c).value)
            if key:
                headers[key] = c

        sku_col = find_header_col(headers, 'sku no', 'sku no.', 'sku')
        sales_col = None
        for key, col in headers.items():
            if key.startswith('stock out'):
                sales_col = col
                break

        if sku_col is None or sales_col is None:
            print(f'  [WARN] Kolom SKU atau Stock Out tidak ditemukan di {path.name}')
            wb.close()
            return sales_data

        for r in range(header_row + 1, ws.max_row + 1):
            sku = ws.cell(r, sku_col).value
            stock_out = safe_float(ws.cell(r, sales_col).value)
            if sku and stock_out > 0:
                sku_str = str(sku).strip()
                sales_data[sku_str] = sales_data.get(sku_str, 0) + stock_out

        wb.close()

    elif ext == '.csv':
        with open(path, encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                sku_key = None
                sales_key = None
                for k in row.keys():
                    k_norm = norm(k)
                    if k_norm in ('sku no', 'sku no.', 'sku'):
                        sku_key = k
                    if k_norm.startswith('stock out'):
                        sales_key = k
                
                if sku_key and sales_key:
                    sku = row.get(sku_key, '').strip()
                    stock_out = safe_float(row.get(sales_key, 0))
                    if sku and stock_out > 0:
                        sales_data[sku] = sales_data.get(sku, 0) + stock_out

    return sales_data


def process_odoo_transfer_to_branch(ws, headers: Dict[str, int], 
                                     odoo_data: List[Dict],
                                     col_name: str,
                                     existing_max_row: int) -> Tuple[int, int]:
    product_col = find_header_col(headers, 'product')
    lot_col = find_header_col(headers, 'lot/serial number')
    target_col = find_header_col(headers, col_name, f'{col_name} ')
    
    if product_col is None or lot_col is None or target_col is None:
        return 0, 0

    updated = 0
    added = 0
    new_row = existing_max_row + 1

    for item in odoo_data:
        product = str(item['product']).strip()
        lot = str(item['lot']).strip()
        qty = item['qty']
        sku = item.get('sku', '')
        sku_batch = item.get('sku_batch', '')

        found = False
        for row in range(3, existing_max_row + 1):
            existing_product = ws.cell(row, product_col).value
            existing_lot = ws.cell(row, lot_col).value
            
            if existing_product and existing_lot:
                if str(existing_product).strip() == product and str(existing_lot).strip() == lot:
                    current = ws.cell(row, target_col).value or 0
                    ws.cell(row, target_col).value = float(current) + qty
                    updated += 1
                    found = True
                    break

        if not found:
            ws.cell(new_row, 1).value = sku
            ws.cell(new_row, 2).value = sku_batch
            ws.cell(new_row, 3).value = item['product']
            ws.cell(new_row, 4).value = lot
            ws.cell(new_row, 7).value = 0
            ws.cell(new_row, 10).value = 0
            ws.cell(new_row, 11).value = 0
            ws.cell(new_row, 13).value = 0
            ws.cell(new_row, target_col).value = qty
            new_row += 1
            added += 1

    return updated, added


def distribute_sales_fefo(ws, ws_val, headers: Dict[str, int],
                          commercial_sales: Dict[str, float]) -> int:
    sku_col = find_header_col(headers, 'sku')
    exp_col = find_header_col(headers, 'expiration date')
    sales_col = find_header_col(headers, 'sales', 'sales ')
    
    if sku_col is None or sales_col is None:
        return 0

    sku_batches: Dict[str, List[Tuple[int, Optional[datetime], float]]] = {}

    for row in range(3, ws.max_row + 1):
        sku = str(first_nonblank(ws_val.cell(row, sku_col).value,
                                 ws.cell(row, sku_col).value, '') or '').strip()
        if not sku:
            continue

        exp_date = None
        if exp_col:
            exp_raw = first_nonblank(ws_val.cell(row, exp_col).value,
                                     ws.cell(row, exp_col).value)
            exp_date = parse_expiry_date(exp_raw)

        if sku not in sku_batches:
            sku_batches[sku] = []
        sku_batches[sku].append((row, exp_date))

    count = 0
    for sku, total_sales in commercial_sales.items():
        sku_upper = sku.strip().upper()
        batches = None
        if sku in sku_batches:
            batches = sku_batches[sku]
        else:
            for k, v in sku_batches.items():
                if k.upper() == sku_upper:
                    batches = v
                    break

        if batches is None:
            continue

        def sort_key(b):
            return b[1] if b[1] is not None else datetime(9999, 12, 31)

        sorted_batches = sorted(batches, key=sort_key)

        remaining_sales = total_sales
        for row, exp_date in sorted_batches:
            if remaining_sales <= 0:
                ws.cell(row, sales_col).value = 0
                continue

            ws.cell(row, sales_col).value = remaining_sales
            remaining_sales = 0
            break

        if remaining_sales > 0 and sorted_batches:
            last_row = sorted_batches[-1][0]
            current = safe_float(ws.cell(last_row, sales_col).value)
            ws.cell(last_row, sales_col).value = current + remaining_sales

        count += 1

    return count


def add_formulas_to_sheet(ws, start_row: int, end_row: int):
    sum_cols = {7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 16: 'P', 17: 'Q'}
    
    for col_num, col_letter in sum_cols.items():
        ws.cell(1, col_num).value = f'=SUM({col_letter}{start_row}:{col_letter}{end_row})'
    
    for row in range(start_row, end_row + 1):
        row_str = str(row)
        ws.cell(row, 1).value = f'=MID(C{row_str},FIND("[",C{row_str})+1,FIND("]",C{row_str})-FIND("[",C{row_str})-1)'
        ws.cell(row, 2).value = f'=A{row_str}&D{row_str}'
        ws.cell(row, 5).value = f'=IF(ISBLANK(F{row_str}),"",DATEDIF(MIN(TODAY(),F{row_str}),MAX(TODAY(),F{row_str}),"M")*IF(F{row_str}>=TODAY(),1,-1))'
        ws.cell(row, 9).value = f'=SUM(G{row_str}:H{row_str})'
        ws.cell(row, 12).value = f'=SUM(I{row_str}:J{row_str})-K{row_str}'
        ws.cell(row, 14).value = f'=L{row_str}-M{row_str}'
        ws.cell(row, 17).value = f'=P{row_str}-M{row_str}'


def main():
    parser = argparse.ArgumentParser(
        description='Buat file laporan stock bulan baru dengan auto-import data.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Contoh penggunaan:
  # Mode otomatis (cocokkan file berdasarkan nama sheet):
  python buat_laporan_stock_baru.py "Stock Feb 2026.xlsx" --folder "C:\\Data\\Mar 2026"
  
  # Mode manual (isi satu per satu):
  python buat_laporan_stock_baru.py "Stock Feb 2026.xlsx" --stock-masuk "Transfer (70).xlsx" --odoo "Transfer (71).xlsx"
  
  # Campuran:
  python buat_laporan_stock_baru.py "Stock Feb 2026.xlsx" --folder "C:\\Data" --commercial "Commercial.xlsx"
        """
    )
    parser.add_argument('sumber', nargs='?', help='Path file sumber (Excel)')
    parser.add_argument('output', nargs='?', default='', help='Path file output (opsional)')
    parser.add_argument('--folder', dest='folder', help='Folder berisi file-file Stock Masuk, Odoo, Commercial')
    parser.add_argument('--stock-masuk', dest='stock_masuk', help='File Odoo Stock Masuk (mode manual)')
    parser.add_argument('--odoo', dest='odoo', help='File Odoo Sales (mode manual)')
    parser.add_argument('--commercial', dest='commercial', help='File Commercial untuk Sales')
    parser.add_argument('--commercial-sheet', dest='commercial_sheet', help='Nama sheet di file Commercial')
    parser.add_argument('--sheets', dest='sheets', nargs='+', help='Sheet yang diproses')

    if len(sys.argv) >= 2 and not sys.argv[1].startswith('-'):
        args = parser.parse_args()
    else:
        args = parser.parse_args([])

    if args.sumber:
        src = Path(args.sumber).expanduser().resolve()
    else:
        src = Path(input('Masukkan path file sumber: ').strip().strip('"')).expanduser().resolve()

    if not src.exists():
        raise FileNotFoundError(f'File tidak ditemukan: {src}')

    if args.output and args.output.strip():
        out = Path(args.output).expanduser().resolve()
    else:
        out = guess_output_name(src)

    # Load workbook to get sheet names
    wb_src = openpyxl.load_workbook(src, read_only=True)
    sheet_names = [s for s in wb_src.sheetnames if s.lower() not in ['total stock', 'inventory at date vs fisik']]
    wb_src.close()

    target_sheets = [s.lower() for s in args.sheets] if args.sheets else None

    print()
    print('=' * 60)
    print('  PEMBUAT FILE LAPORAN STOCK BULAN BARU')
    print('=' * 60)
    print(f'  Sumber  : {src.name}')
    print(f'  Output  : {out.name}')
    
    # Prepare file matching
    folder_path = Path(args.folder).expanduser().resolve() if args.folder else None
    sheet_files = {}
    
    if folder_path and folder_path.exists():
        print(f'  Folder  : {folder_path}')
        sheet_files = match_files_to_sheets(folder_path, sheet_names)
        print()
        print('  File yang ditemukan:')
        for sheet_name, files in sheet_files.items():
            stock_masuk_file = files['stock_masuk'].name if files['stock_masuk'] else '-'
            odoo_file = files['odoo'].name if files['odoo'] else '-'
            print(f'    {sheet_name}:')
            print(f'      Stock Masuk: {stock_masuk_file}')
            print(f'      Odoo:        {odoo_file}')
    else:
        # Manual mode - use same files for all sheets
        if args.stock_masuk:
            for sheet_name in sheet_names:
                sheet_files[sheet_name] = {'stock_masuk': Path(args.stock_masuk).expanduser().resolve(), 'odoo': None, 'commercial': None}
        if args.odoo:
            for sheet_name in sheet_names:
                if sheet_name in sheet_files:
                    sheet_files[sheet_name]['odoo'] = Path(args.odoo).expanduser().resolve()
                else:
                    sheet_files[sheet_name] = {'stock_masuk': None, 'odoo': Path(args.odoo).expanduser().resolve(), 'commercial': None}
    
    if folder_path:
        print(f'  Sheets  : {", ".join([s for s in sheet_names if target_sheets is None or s.lower() in target_sheets])[:50]}...' if len(sheet_names) > 5 else f'  Sheets  : {", ".join(sheet_names)}')
    else:
        print(f'  Sheets  : SEMUA')
    print('=' * 60)
    print()

    # Load commercial data (shared across sheets if from folder)
    commercial_path = Path(args.commercial).expanduser().resolve() if args.commercial else None
    commercial_sales_data = {}
    commercial_sheet = args.commercial_sheet

    if commercial_path and commercial_path.exists():
        print(f'Membaca file Commercial: {commercial_path.name}...')
        commercial_sales_data = load_commercial_data(commercial_path, commercial_sheet)
        print(f'  → {len(commercial_sales_data)} SKU Sales ditemukan')
    else:
        print('File Commercial: tidak ada (opsional, dilewati)')

    print()

    print(f'Membaca file sumber: {src.name}...')
    wb = openpyxl.load_workbook(src, data_only=False)
    wb_values = openpyxl.load_workbook(src, data_only=True)

    stock_map: Dict[Tuple[str, str, str], float] = {}

    summary = {
        'sheets_processed': 0,
        'rows_reset': 0,
        'stock_masuk_filled': 0,
        'odoo_filled': 0,
        'sales_distributed': 0,
        'errors': [],
    }

    for sheet_name in wb.sheetnames:
        if target_sheets and sheet_name.lower() not in target_sheets:
            continue

        ws = wb[sheet_name]
        if not is_branch_sheet(ws):
            continue

        print(f'Memproses sheet: "{sheet_name}"...')
        summary['sheets_processed'] += 1

        try:
            ws_val = wb_values[sheet_name]
            headers = header_map(ws)

            stock_awal_col = find_header_col(headers, 'stock awal', 'stock awal ')
            stock_akhir_col = find_header_col(headers, 'stock akhir', 'stock akhir ')
            sku_col = find_header_col(headers, 'sku')
            lot_col = find_header_col(headers, 'lot/serial number')

            if stock_awal_col is None or stock_akhir_col is None:
                print(f'  [WARN] Kolom Stock Awal / Stock Akhir tidak ditemukan, skip.')
                continue

            existing_max_row = ws.max_row

            for row in range(3, ws.max_row + 1):
                row_has_data = any(
                    ws.cell(row, c).value not in (None, '')
                    for c in range(1, min(ws.max_column, 14) + 1)
                )
                if not row_has_data:
                    continue

                closing_value = safe_float(ws_val.cell(row, stock_akhir_col).value)
                ws.cell(row, stock_awal_col).value = closing_value

                for col_name in RESET_COLUMNS:
                    col_idx = headers.get(col_name)
                    if col_idx:
                        ws.cell(row, col_idx).value = 0

                summary['rows_reset'] += 1

                sku_val = None
                lot_val = None
                if sku_col:
                    sku_val = first_nonblank(ws_val.cell(row, sku_col).value,
                                             ws.cell(row, sku_col).value)
                if lot_col:
                    lot_val = first_nonblank(ws_val.cell(row, lot_col).value,
                                             ws.cell(row, lot_col).value)

                location = f'{sheet_name}/PAYU'
                if sku_val not in (None, '') and lot_val not in (None, ''):
                    stock_map[(str(location), str(sku_val), str(lot_val))] = closing_value

            # Get files for this sheet
            sheet_file = sheet_files.get(sheet_name, {})
            stock_masuk_path = sheet_file.get('stock_masuk')
            odoo_path = sheet_file.get('odoo')

            # Load and process Stock Masuk
            stock_masuk_data = []
            if stock_masuk_path and stock_masuk_path.exists():
                stock_masuk_data = load_odoo_transfer_data(stock_masuk_path)
                updated, added = process_odoo_transfer_to_branch(ws, headers, stock_masuk_data, 'stock masuk', existing_max_row)
                summary['stock_masuk_filled'] += updated + added
                existing_max_row = ws.max_row
                if updated or added:
                    print(f'  → Stock Masuk ({stock_masuk_path.name}): {updated} diupdate, {added} baru ditambahkan')

            # Load and process Odoo
            odoo_data = []
            if odoo_path and odoo_path.exists():
                odoo_data = load_odoo_transfer_data(odoo_path)
                updated, added = process_odoo_transfer_to_branch(ws, headers, odoo_data, 'odoo', existing_max_row)
                summary['odoo_filled'] += updated + added
                existing_max_row = ws.max_row
                if updated or added:
                    print(f'  → Odoo ({odoo_path.name}): {updated} diupdate, {added} baru ditambahkan')

            # Distribute Sales from Commercial (FEFO)
            if commercial_sales_data:
                distributed = distribute_sales_fefo(ws, ws_val, headers, commercial_sales_data)
                summary['sales_distributed'] += distributed
                if distributed:
                    print(f'  → Sales FEFO: {distributed} SKU didistribusi')

            new_max_row = ws.max_row
            add_formulas_to_sheet(ws, 3, new_max_row)

        except Exception as e:
            error_msg = f'Error pada sheet "{sheet_name}": {e}'
            print(f'  [ERROR] {error_msg}')
            summary['errors'].append(error_msg)

    ws_total = find_sheet_ci(wb, 'Total Stock')
    if ws_total:
        print('Memproses sheet: "Total Stock"...')
        headers = header_map(ws_total)
        sku_col = headers.get('sku')
        location_col = headers.get('location')
        lot_col = headers.get('lot/serial number')
        stock_col = find_header_col(headers, 'stock akhir', 'stock akhir ')
        if sku_col and location_col and lot_col and stock_col:
            for row in range(3, ws_total.max_row + 1):
                key = (
                    str(first_nonblank(ws_total.cell(row, location_col).value, '')),
                    str(first_nonblank(ws_total.cell(row, sku_col).value, '')),
                    str(first_nonblank(ws_total.cell(row, lot_col).value, '')),
                )
                if key in stock_map:
                    ws_total.cell(row, stock_col).value = stock_map[key]

    ws_inv = find_sheet_ci(wb, 'Inventory AT Date vs Fisik')
    if ws_inv:
        print('Memproses sheet: "Inventory AT Date vs Fisik"...')
        headers = header_map(ws_inv)
        sku_col = headers.get('sku')
        location_col = headers.get('location')
        lot_col = headers.get('lot/serial number')
        qoh_col = headers.get('quantity on hand')
        fisik_col = headers.get('stock fisik')
        ket_col = headers.get('keterangan')
        if sku_col and location_col and lot_col and qoh_col and fisik_col:
            for row in range(3, ws_inv.max_row + 1):
                key = (
                    str(first_nonblank(ws_inv.cell(row, location_col).value, '')),
                    str(first_nonblank(ws_inv.cell(row, sku_col).value, '')),
                    str(first_nonblank(ws_inv.cell(row, lot_col).value, '')),
                )
                if key in stock_map:
                    value = stock_map[key]
                    ws_inv.cell(row, qoh_col).value = value
                    ws_inv.cell(row, fisik_col).value = value
                    if ket_col:
                        ws_inv.cell(row, ket_col).value = None

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcOnSave = True
    except Exception:
        pass

    wb.save(out)

    print()
    print('=' * 60)
    print('  RINGKASAN')
    print('=' * 60)
    print(f'  Sheet diproses      : {summary["sheets_processed"]}')
    print(f'  Baris di-reset      : {summary["rows_reset"]}')
    if summary['stock_masuk_filled'] > 0:
        print(f'  Stock Masuk terisi  : {summary["stock_masuk_filled"]}')
    if summary['odoo_filled'] > 0:
        print(f'  Odoo terisi         : {summary["odoo_filled"]}')
    if commercial_sales_data:
        print(f'  Sales FEFO (SKU)    : {summary["sales_distributed"]}')
    if summary['errors']:
        print(f'  Errors              : {len(summary["errors"])}')
        for err in summary['errors']:
            print(f'    - {err}')
    print(f'\n  File berhasil dibuat: {out}')
    print('=' * 60)


if __name__ == '__main__':
    main()